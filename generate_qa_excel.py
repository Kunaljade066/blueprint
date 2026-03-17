#!/usr/bin/env python3
"""
QA Blueprint — Professional Excel Test Case Generator
Meets all formatting requirements for stakeholder-ready output.
Usage:
    python3 generate_qa_excel.py input.json output.xlsx
    OR called from qa-blueprint HTML via JSON paste
"""

import json, sys, os, io
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


# ──────────────────────────────────────────────
#  COLOUR PALETTE
# ──────────────────────────────────────────────
C = {
    # Header
    "hdr_bg":        "1F3864",   # dark navy blue
    "hdr_fg":        "FFFFFF",   # white

    # Title banner
    "title_bg":      "0D0D1A",
    "title_fg":      "FFFFFF",

    # Priority — row background (light tints)
    "p0_row":        "FFE8E8",   # soft red
    "p1_row":        "FFF0D9",   # soft orange
    "p2_row":        "FFFDE7",   # soft yellow
    "p3_row":        "E8F5E9",   # soft green

    # Priority — cell badge (saturated)
    "p0_cell":       "C0392B",   # red text
    "p0_cell_bg":    "FDDCDC",
    "p1_cell":       "D35400",   # orange text
    "p1_cell_bg":    "FDEBD0",
    "p2_cell":       "B7950B",   # dark yellow text
    "p2_cell_bg":    "FEF9E7",
    "p3_cell":       "1E8449",   # green text
    "p3_cell_bg":    "D5F5E3",

    # Category tints
    "cat_functional":  "D6EAF8",
    "cat_regression":  "E8DAED",
    "cat_edge":        "FEF9E7",
    "cat_ui":          "D5F5E3",
    "cat_perf":        "FDEDEC",
    "cat_integ":       "DDEEFF",
    "cat_security":    "FDF2F8",
    "cat_default":     "EAF2FF",

    # Status
    "st_notstarted_bg": "F2F3F4",  "st_notstarted_fg": "7F8C8D",
    "st_inprogress_bg": "D6EAF8",  "st_inprogress_fg": "1A5276",
    "st_pass_bg":       "D5F5E3",  "st_pass_fg":       "1E8449",
    "st_fail_bg":       "FDEDEC",  "st_fail_fg":       "C0392B",
    "st_blocked_bg":    "FEF9E7",  "st_blocked_fg":    "D35400",

    # General
    "zebra_even":    "F7F9FC",
    "zebra_odd":     "FFFFFF",
    "border_thin":   "C5CAE9",
    "border_hdr":    "3D8EFF",
    "tc_id_fg":      "1565C0",
    "tc_id_bg_e":    "E3F2FD",
    "tc_id_bg_o":    "EEF6FD",
    "steps_bg_e":    "F0F6FF",
    "steps_bg_o":    "E8F2FF",
    "exp_bg_e":      "F0FBF5",
    "exp_bg_o":      "E8F8EE",
    "bp_fg":         "C0392B",
    "bp_bg_e":       "FEF5F5",
    "bp_bg_o":       "FDF0F0",
    "precon_bg_e":   "FAFBFC",
    "precon_bg_o":   "F5F6F8",
    "sys_bg_e":      "F8F9FA",
    "sys_bg_o":      "F2F3F4",
}

# ──────────────────────────────────────────────
#  HELPERS
# ──────────────────────────────────────────────
def fl(hex6): return PatternFill("solid", fgColor=hex6)
def fn(color="000000", size=10, bold=False, name="Calibri"):
    return Font(name=name, size=size, bold=bold, color=color)
def sd(style="thin", color="C5CAE9"): return Side(style=style, color=color)

THIN   = Border(left=sd(), right=sd(), top=sd(), bottom=sd())
MEDIUM = Border(
    left=  sd("medium", C["border_hdr"]),
    right= sd("medium", C["border_hdr"]),
    top=   sd("medium", C["border_hdr"]),
    bottom=sd("medium", C["border_hdr"])
)

def al(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

# Priority → row tint
PRIO_ROW = {
    "P0":"p0_row","P1":"p1_row","P2":"p2_row","P3":"p3_row",
    "Critical":"p0_row","High":"p1_row","Medium":"p2_row","Low":"p3_row",
}
# Priority → badge colours
PRIO_BADGE = {
    "P0":       (C["p0_cell"],    C["p0_cell_bg"]),
    "P1":       (C["p1_cell"],    C["p1_cell_bg"]),
    "P2":       (C["p2_cell"],    C["p2_cell_bg"]),
    "P3":       (C["p3_cell"],    C["p3_cell_bg"]),
    "Critical": (C["p0_cell"],    C["p0_cell_bg"]),
    "High":     (C["p1_cell"],    C["p1_cell_bg"]),
    "Medium":   (C["p2_cell"],    C["p2_cell_bg"]),
    "Low":      (C["p3_cell"],    C["p3_cell_bg"]),
}
# Category → tint
CAT_FILL = {
    "Functional":  C["cat_functional"],
    "Regression":  C["cat_regression"],
    "Edge Case":   C["cat_edge"],
    "UI":          C["cat_ui"],
    "Performance": C["cat_perf"],
    "Integration": C["cat_integ"],
    "Security":    C["cat_security"],
}
# Status badge
STATUS_BADGE = {
    "Not Started": (C["st_notstarted_fg"], C["st_notstarted_bg"]),
    "In Progress": (C["st_inprogress_fg"], C["st_inprogress_bg"]),
    "Pass":        (C["st_pass_fg"],       C["st_pass_bg"]),
    "Passed":      (C["st_pass_fg"],       C["st_pass_bg"]),
    "Fail":        (C["st_fail_fg"],       C["st_fail_bg"]),
    "Failed":      (C["st_fail_fg"],       C["st_fail_bg"]),
    "Blocked":     (C["st_blocked_fg"],    C["st_blocked_bg"]),
}

HEADERS = [
    "TC ID", "Feature", "Test Title", "Category",
    "Preconditions", "Steps", "Expected Result",
    "Break Point", "Systems", "Edge Case",
    "Priority", "Status", "Dev Result", "QA Result"
]
# (width_chars, h_align, wrap)
COL_CONFIG = [
    (10,  "center", False),  # TC ID
    (18,  "left",   False),  # Feature
    (44,  "left",   True),   # Test Title
    (14,  "center", False),  # Category
    (32,  "left",   True),   # Preconditions
    (46,  "left",   True),   # Steps
    (46,  "left",   True),   # Expected Result
    (30,  "left",   True),   # Break Point
    (18,  "center", True),   # Systems
    (11,  "center", False),  # Edge Case
    (11,  "center", False),  # Priority
    (13,  "center", False),  # Status
    (13,  "center", False),  # Dev Result
    (13,  "center", False),  # QA Result
]


# ──────────────────────────────────────────────
#  MAIN GENERATOR
# ──────────────────────────────────────────────
def generate(tcs: list, feature_name: str = "Test Cases", out_path: str = None):
    wb = Workbook()

    # ── SHEET 1: Test Cases ──────────────────
    ws = wb.active
    ws.title = feature_name[:31]
    ws.sheet_properties.tabColor = "1F3864"

    N = len(HEADERS)
    last_col = get_column_letter(N)

    # ── Row 1: Title Banner ──
    ws.merge_cells(f"A1:{last_col}1")
    tc = ws["A1"]
    tc.value     = f"  QA Blueprint · Hitwicket  —  {feature_name}"
    tc.font      = fn(C["title_fg"], 13, True, "Calibri")
    tc.fill      = fl(C["title_bg"])
    tc.alignment = al("left", "center")
    ws.row_dimensions[1].height = 38

    # ── Row 2: Column Headers ──
    for ci, h in enumerate(HEADERS, 1):
        cell = ws.cell(row=2, column=ci, value=h)
        cell.font      = fn(C["hdr_fg"], 10, True, "Calibri")
        cell.fill      = fl(C["hdr_bg"])
        cell.alignment = al("center", "center", True)
        cell.border    = MEDIUM
    ws.row_dimensions[2].height = 32

    # ── Rows 3+: Data ──
    for ri, tc_data in enumerate(tcs, 3):
        prio   = str(tc_data.get("priority", "")).strip()
        cat    = str(tc_data.get("category", "")).strip()
        status = str(tc_data.get("status", "Not Started")).strip() or "Not Started"
        edge   = str(tc_data.get("edge_case") or tc_data.get("is_edge_case") or "").strip()

        # Row tint based on priority
        row_bg = C.get(PRIO_ROW.get(prio, ""), C["zebra_even" if ri%2==0 else "zebra_odd"])

        steps_raw = tc_data.get("steps", "") or ""
        steps_val = steps_raw.replace("\\n", "\n")

        row_vals = [
            tc_data.get("id", ""),
            tc_data.get("feature", ""),
            tc_data.get("title") or tc_data.get("test_title", ""),
            cat,
            tc_data.get("preconditions", ""),
            steps_val,
            tc_data.get("expected") or tc_data.get("expected_result", ""),
            tc_data.get("break_point") or tc_data.get("breakpoint", ""),
            tc_data.get("systems") or tc_data.get("affected_systems", ""),
            edge,
            prio,
            status,
            tc_data.get("dev_result", ""),
            tc_data.get("qa_result", ""),
        ]

        for ci, val in enumerate(row_vals, 1):
            cell = ws.cell(row=ri, column=ci, value=str(val) if val else "")
            cell.border = THIN
            cfg_w, cfg_h, cfg_wrap = COL_CONFIG[ci - 1]

            # ── Per-column styling ──
            if ci == 1:  # TC ID
                bg = C["tc_id_bg_e"] if ri % 2 == 0 else C["tc_id_bg_o"]
                cell.font      = fn(C["tc_id_fg"], 9, True, "Courier New")
                cell.fill      = fl(bg)
                cell.alignment = al("center", "center")

            elif ci == 2:  # Feature
                cell.font      = fn("2C3E50", 10, True)
                cell.fill      = fl(row_bg)
                cell.alignment = al("left", "center", False)

            elif ci == 3:  # Test Title
                cell.font      = fn("0D0D1A", 10, True)
                cell.fill      = fl(row_bg)
                cell.alignment = al("left", "center", True)

            elif ci == 4:  # Category
                cat_bg = CAT_FILL.get(cat, C["cat_default"])
                cell.font      = fn("2C3E50", 9)
                cell.fill      = fl(cat_bg)
                cell.alignment = al("center", "center")

            elif ci == 5:  # Preconditions
                bg = C["precon_bg_e"] if ri % 2 == 0 else C["precon_bg_o"]
                cell.font      = fn("5D6D7E", 9)
                cell.fill      = fl(bg)
                cell.alignment = al("left", "top", True)

            elif ci == 6:  # Steps
                bg = C["steps_bg_e"] if ri % 2 == 0 else C["steps_bg_o"]
                cell.font      = fn("1A1A2E", 9)
                cell.fill      = fl(bg)
                cell.alignment = al("left", "top", True)

            elif ci == 7:  # Expected Result
                bg = C["exp_bg_e"] if ri % 2 == 0 else C["exp_bg_o"]
                cell.font      = fn("145A32", 9)
                cell.fill      = fl(bg)
                cell.alignment = al("left", "top", True)

            elif ci == 8:  # Break Point
                bg = C["bp_bg_e"] if ri % 2 == 0 else C["bp_bg_o"]
                cell.font      = fn(C["bp_fg"], 9)
                cell.fill      = fl(bg)
                cell.alignment = al("left", "top", True)

            elif ci == 9:  # Systems
                bg = C["sys_bg_e"] if ri % 2 == 0 else C["sys_bg_o"]
                cell.font      = fn("5D6D7E", 8)
                cell.fill      = fl(bg)
                cell.alignment = al("center", "center", True)

            elif ci == 10:  # Edge Case
                is_yes = val.lower() == "yes"
                fg = "CA6F1E" if is_yes else "7F8C8D"
                bg = "FEF5E7" if is_yes else row_bg
                cell.font      = fn(fg, 9, is_yes)
                cell.fill      = fl(bg)
                cell.alignment = al("center", "center")

            elif ci == 11:  # Priority
                pfg, pbg = PRIO_BADGE.get(prio, ("555555", "EEEEEE"))
                cell.font      = fn(pfg, 9, True)
                cell.fill      = fl(pbg)
                cell.alignment = al("center", "center")

            elif ci == 12:  # Status
                sfg, sbg = STATUS_BADGE.get(status, (C["st_notstarted_fg"], C["st_notstarted_bg"]))
                cell.font      = fn(sfg, 9)
                cell.fill      = fl(sbg)
                cell.alignment = al("center", "center")

            else:  # Dev Result, QA Result
                cell.font      = fn("7F8C8D", 9)
                cell.fill      = fl("F8F9FA")
                cell.alignment = al("center", "center")

        ws.row_dimensions[ri].height = 75

    # ── Column widths ──
    for ci, (w, _, _) in enumerate(COL_CONFIG, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    # ── Freeze below title + header (row 3) ──
    ws.freeze_panes = "A3"

    # ── Auto-filter on header row ──
    ws.auto_filter.ref = f"A2:{last_col}2"

    # ── Print settings (nice when printing) ──
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToPage   = True
    ws.page_setup.fitToWidth  = 1
    ws.page_margins.left       = 0.5
    ws.page_margins.right      = 0.5
    ws.page_margins.top        = 0.75
    ws.page_margins.bottom     = 0.75

    # ──────────────────────────────────────────
    #  SHEET 2: Summary Dashboard
    # ──────────────────────────────────────────
    ws2 = wb.create_sheet("📊 Summary")
    ws2.sheet_properties.tabColor = "00D4A0"
    ws2.column_dimensions["A"].width = 28
    ws2.column_dimensions["B"].width = 16
    ws2.column_dimensions["C"].width = 20

    # Title
    ws2.merge_cells("A1:C1")
    t2 = ws2["A1"]
    t2.value     = "  Test Suite Summary Dashboard"
    t2.font      = fn("FFFFFF", 14, True)
    t2.fill      = fl(C["title_bg"])
    t2.alignment = al("left", "center")
    ws2.row_dimensions[1].height = 42

    # Subtitle
    ws2.merge_cells("A2:C2")
    s2 = ws2["A2"]
    s2.value     = f"  {feature_name}  ·  {len(tcs)} Test Cases"
    s2.font      = fn("AAAACC", 10)
    s2.fill      = fl("1A1A2E")
    s2.alignment = al("left", "center")
    ws2.row_dimensions[2].height = 24

    # Gap
    ws2.row_dimensions[3].height = 10

    # Sub-header row
    for ci, h in enumerate(["METRIC", "COUNT", "STATUS"], 1):
        c2 = ws2.cell(row=4, column=ci, value=h)
        c2.font      = fn("FFFFFF", 10, True)
        c2.fill      = fl(C["hdr_bg"])
        c2.alignment = al("center", "center")
        c2.border    = MEDIUM
    ws2.row_dimensions[4].height = 28

    # Counts
    p = lambda k: sum(1 for t in tcs if t.get("priority","") == k)
    ec = sum(1 for t in tcs if str(t.get("edge_case","")).lower() == "yes")
    tot = len(tcs)

    summary_rows = [
        ("Total Test Cases",     tot,                    "All",         "2471A3", "E8F4FD"),
        ("──────────────────",   "",                     "",            "CCCCCC", "F8F9FA"),
        ("P0 — Blocker",         p("P0")+p("Critical"),  "Critical",    C["p0_cell"],C["p0_cell_bg"]),
        ("P1 — Critical",        p("P1")+p("High"),      "High",        C["p1_cell"],C["p1_cell_bg"]),
        ("P2 — Important",       p("P2")+p("Medium"),    "Medium",      C["p2_cell"],C["p2_cell_bg"]),
        ("P3 — Low",             p("P3")+p("Low"),       "Low",         C["p3_cell"],C["p3_cell_bg"]),
        ("──────────────────",   "",                     "",            "CCCCCC", "F8F9FA"),
        ("Edge Cases",           ec,                     "Watch",       "CA6F1E", "FEF5E7"),
        ("Automated Eligible",   sum(1 for t in tcs if t.get("category","") in ("Regression","Functional")), "Automate", "1565C0","E3F2FD"),
        ("──────────────────",   "",                     "",            "CCCCCC", "F8F9FA"),
        ("Not Started",          tot,                    "Pending",     C["st_notstarted_fg"],C["st_notstarted_bg"]),
        ("In Progress",          0,                      "Active",      C["st_inprogress_fg"],C["st_inprogress_bg"]),
        ("Passed",               0,                      "Done",        C["st_pass_fg"],       C["st_pass_bg"]),
        ("Failed",               0,                      "Action Req.", C["st_fail_fg"],       C["st_fail_bg"]),
    ]

    for ri2, (label, count, badge, fg, bg) in enumerate(summary_rows, 5):
        ra = ws2.cell(row=ri2, column=1, value=label)
        rb = ws2.cell(row=ri2, column=2, value=count if count != "" else "")
        rc = ws2.cell(row=ri2, column=3, value=badge)
        ws2.row_dimensions[ri2].height = 30

        if label.startswith("──"):
            for c2 in [ra, rb, rc]:
                c2.fill = fl("F0F2F5")
            continue

        ra.font      = fn("2C3E50", 10)
        ra.fill      = fl("F8F9FA")
        ra.alignment = al("left", "center", indent=2)
        ra.border    = THIN

        rb.font      = fn(fg, 14, True)
        rb.fill      = fl(bg)
        rb.alignment = al("center", "center")
        rb.border    = THIN

        rc.font      = fn(fg, 9, False)
        rc.fill      = fl(bg)
        rc.alignment = al("center", "center")
        rc.border    = THIN

    # ── Save ──
    if out_path:
        wb.save(out_path)
        return out_path
    else:
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf.read()


# ──────────────────────────────────────────────
#  CLI ENTRY POINT
# ──────────────────────────────────────────────
if __name__ == "__main__":
    if len(sys.argv) > 1 and os.path.isfile(sys.argv[1]):
        with open(sys.argv[1]) as f:
            data = json.load(f)
    else:
        print("Reading JSON from stdin...")
        data = json.load(sys.stdin)

    tcs          = data.get("test_cases", data) if isinstance(data, dict) else data
    feature_name = data.get("feature_name", "Test Cases") if isinstance(data, dict) else "Test Cases"
    out          = sys.argv[2] if len(sys.argv) > 2 else f"{feature_name.replace(' ','_')}_QA_TCs.xlsx"

    result = generate(tcs, feature_name, out)
    print(f"✓  Saved → {result}")
    print(f"   {len(tcs)} test cases  |  {len(tcs)} rows  |  14 columns")
